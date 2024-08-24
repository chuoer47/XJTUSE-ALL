from wordcloud import WordCloud
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# 打开xlsx
from wordcloud import WordCloud

workbook = load_workbook('西安二手房_清洗版.xlsx')
worksheet = workbook['Sheet 1']

'''
# 把地址写入txt文件中
data = []
f = open("地址.txt", 'w', encoding="gbk")
for row in range(2, worksheet.max_row+1):
    data.append(worksheet['B'+str(row)].value)
f.write(str(data))
f.close()

# 读取
f = open("地址.txt", 'r', encoding="gbk")
outstr = f.readline().strip()
'''

# 生成词云图
outstr = ""
for row in range(2, worksheet.max_row+1):
    outstr += worksheet['B'+str(row)].value+" "
font = r'C:\Windows\Fonts\simsun.ttc'
wc = WordCloud(font_path=font, width=2400, height=1200, max_words=100)
wc.generate(outstr)
wc.to_file('词云.jpg')
plt.figure(dpi=100)
plt.imshow(wc, interpolation='catrom')
plt.axis('off')
plt.show()
plt.close()