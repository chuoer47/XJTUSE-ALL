from openpyxl import load_workbook

# 打开xlsx
workbook = load_workbook('西安二手房_清洗版.xlsx')
worksheet = workbook['Sheet 1']

# 把地址写入txt文件中
data = []
for row in range(2, worksheet.max_row + 1):
    data.append(worksheet['B' + str(row)].value)
name = set(data)
res = []
for n in name:
    sum = 0
    for i in data:
        if i == n:
            sum += 1
    res.append((n, sum / len(data)))
res.sort(key=lambda res: res[1], reverse=True)
for name, fre in res:
    print('%s %0.2f%%' % (name, fre * 100))
