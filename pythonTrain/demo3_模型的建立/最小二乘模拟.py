import numpy as np
from scipy import optimize
from openpyxl import load_workbook

workbook = load_workbook('西安二手房_清洗版_版本4_数字编码.xlsx')
worksheet = workbook['编码数据_版本4_数字编码']

# 把数据写入xi，y变量中
x = [[] for i in range(6)]
for row in range(2, worksheet.max_row + 1):
    for name in ['A', 'B', 'C', 'D', 'E', 'F']:
        x[ord(name) - ord('A')].append(worksheet[name + str(row)].value)
y = []
for row in range(2, worksheet.max_row + 1):
    y.append(worksheet['G' + str(row)].value)
for i in range(len(x)):
    x[i] = np.array(x[i])
y = np.array(y)


def residuals(p):
    k0, k1, k2, k3, k4, k5, b = p
    return y - (k0 * x[0] + k1 * x[1] + k2 * x[2] + k3 * x[3] + k4 * x[4] + k4 * x[5] + b)


r = optimize.leastsq(residuals, [1, 1, 1, 1, 1, 1, 0])
print(r[0])
